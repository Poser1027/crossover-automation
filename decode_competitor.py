"""
Competitor Model Code Decoder
==============================

Decode a competitor's orbital motor model code into:
  - Raw decoded fields (per position in the part number)
  - A normalised spec dict ready to feed into spec_matching.find_matches()

Currently supports:
  - Char-Lynn (Eaton) 2000 Series — 30-digit code starting M02...
  - Char-Lynn (Eaton) T Series   — 25-digit code starting MT0...

Looks up codes from competitor_code_extractor.xlsx (must sit next to the script).
New competitor series can be added simply by adding a new sheet to the workbook
in the same layout (model code template + lookup sections).

USAGE
-----
    # CLI
    python decode_competitor.py "M02 2 080 AB 01 AA 00 0 00 1 0 00 00 00 AA AA F"
    python decode_competitor.py M02-2-080-AB-01-AA-00-0-00-1-0-00-00-00-AA-AA-F
    python decode_competitor.py                  # interactive prompt

    # Programmatic
    from decode_competitor import decode_competitor_code, print_decoded
    result = decode_competitor_code("M02 2 080 AB 01 AA 00 0 00 1 0 00 00 00 AA AA F")
    print_decoded(result)
    # Then feed result["specs_for_crossover"] to spec_matching.find_matches()

REQUIREMENTS
------------
- Python 3.9+
- openpyxl
- competitor_code_extractor.xlsx in the same folder
"""

from __future__ import annotations
from dataclasses import dataclass, field
from pathlib import Path
from typing import Optional
from openpyxl import load_workbook
import re
import sys


# ─────────────────────────────────────────────────────────────────────────────
# Series identification — maps part-number prefix to workbook sheet
# ─────────────────────────────────────────────────────────────────────────────
SERIES_PREFIXES = [
    # (prefix, brand, series, sheet_name)
    ("M02", "Char-Lynn (Eaton)", "2000 Series", "CharLynn_2000"),
    ("MT0", "Char-Lynn (Eaton)", "T Series",    "CharLynn_T"),
]


# ─────────────────────────────────────────────────────────────────────────────
# Sheet-layout cache
# ─────────────────────────────────────────────────────────────────────────────
@dataclass
class PositionGroup:
    """One row from the MODEL CODE TEMPLATE."""
    positions: str          # e.g. "4-6", "15", "1-2"
    group: str              # e.g. "Displacement"
    example: str
    defines: str
    lookup_section: str     # e.g. "POS 4-6 — Displacement" or "(fixed)"
    width: int              # character count of this group


@dataclass
class SheetLayout:
    """Parsed model-code structure for one competitor series."""
    sheet_name: str
    brand: str
    series: str
    position_groups: list[PositionGroup]
    lookups: dict[str, dict[str, dict]]   # section_title -> code -> row data


def _detect_series(raw: str) -> Optional[tuple[str, str, str, str]]:
    """Find the matching series tuple for a raw input string."""
    s = re.sub(r"[\s\-/.]", "", raw).upper()
    for prefix, brand, series, sheet in SERIES_PREFIXES:
        if s.startswith(prefix):
            return prefix, brand, series, sheet
    return None


def _parse_sheet(wb, sheet_name: str, brand: str, series: str) -> SheetLayout:
    """Parse a competitor sheet into a SheetLayout object."""
    ws = wb[sheet_name]

    # ─── Find the MODEL CODE TEMPLATE section ───
    template_header_row = None
    for r in range(1, ws.max_row + 1):
        v = ws.cell(row=r, column=1).value
        if v and isinstance(v, str) and "MODEL CODE TEMPLATE" in v.upper():
            template_header_row = r
            break

    if template_header_row is None:
        raise ValueError(f"Could not find MODEL CODE TEMPLATE in sheet '{sheet_name}'")

    # The next row is the column-header row; data starts row+2
    position_groups: list[PositionGroup] = []
    r = template_header_row + 2
    while r <= ws.max_row:
        positions_val = ws.cell(row=r, column=1).value
        group_val = ws.cell(row=r, column=2).value
        if not positions_val or not group_val:
            break
        if str(positions_val).strip().upper() in ("EXAMPLE PART NUMBER", "POSITIONS"):
            break
        example = ws.cell(row=r, column=3).value or ""
        defines = ws.cell(row=r, column=4).value or ""
        lookup_section = ws.cell(row=r, column=5).value or ""
        width = _position_width(str(positions_val), str(example))
        position_groups.append(PositionGroup(
            positions=str(positions_val).strip(),
            group=str(group_val).strip(),
            example=str(example).strip(),
            defines=str(defines).strip(),
            lookup_section=str(lookup_section).strip(),
            width=width,
        ))
        r += 1

    # ─── Find all lookup sections (rows whose col-A starts with "POS ") ───
    lookups: dict[str, dict[str, dict]] = {}
    for r in range(1, ws.max_row + 1):
        v = ws.cell(row=r, column=1).value
        if not (v and isinstance(v, str)):
            continue
        v = v.strip()
        if not v.startswith("POS "):
            continue
        # This is a section title row. The header row is r+1, data starts r+2.
        section_title = v
        header_row = r + 1
        # Detect column count for this section
        n_cols = 0
        for c in range(1, 11):
            if ws.cell(row=header_row, column=c).value:
                n_cols = c
            else:
                break
        col_names = [ws.cell(row=header_row, column=c).value for c in range(1, n_cols + 1)]

        # Read data rows until we hit a blank or another section title
        section_data: dict[str, dict] = {}
        dr = header_row + 1
        while dr <= ws.max_row:
            code_val = ws.cell(row=dr, column=1).value
            if code_val is None or code_val == "":
                break
            code_str = str(code_val).strip()
            # Stop if we've hit the next section title
            if code_str.startswith("POS "):
                break
            row_data = {}
            for c in range(2, n_cols + 1):
                key = col_names[c - 1] or f"col_{c}"
                row_data[str(key)] = ws.cell(row=dr, column=c).value
            section_data[code_str] = row_data
            dr += 1

        lookups[section_title] = section_data

    return SheetLayout(
        sheet_name=sheet_name, brand=brand, series=series,
        position_groups=position_groups, lookups=lookups,
    )


def _position_width(positions: str, example: str = "") -> int:
    """
    Return the character width of a 'positions' field.

    Catalogs number positions inconsistently (e.g. 2000 Series shows "1-2" as a
    single position group whose example value 'M02' is actually 3 chars).
    The example value is the most reliable source of width.
    """
    # Prefer the example's length when available (most reliable)
    if example and example.strip():
        return len(example.strip())

    # Fall back to parsing the position numbers
    s = positions.strip()
    if "-" in s:
        a, b = s.split("-", 1)
        try:
            return int(b) - int(a) + 1
        except ValueError:
            return len(s)
    try:
        int(s)
        return 1
    except ValueError:
        return len(s)


# ─────────────────────────────────────────────────────────────────────────────
# Tokenisation
# ─────────────────────────────────────────────────────────────────────────────
def _tokenise(raw: str, layout: SheetLayout) -> tuple[list[str], list[str]]:
    """
    Split a raw model code into one token per position group.
    Handles dash, space, slash, dot, or no separator.
    Returns (tokens, warnings).
    """
    warnings = []
    s = raw.strip().upper()

    # Try separator-based split first
    tokens_by_sep = [t for t in re.split(r"[\s\-/.]+", s) if t]

    expected_widths = [pg.width for pg in layout.position_groups]
    expected_count = len(expected_widths)

    # ─── Strategy A: separator split matches expected count ───
    if len(tokens_by_sep) == expected_count:
        return tokens_by_sep, warnings

    # ─── Strategy B: separator split has more tokens than groups ───
    # Some users separate every char (M T 0 0 8 0 ...). Reassemble.
    if len(tokens_by_sep) > expected_count:
        compact = "".join(tokens_by_sep)
        return _slice_by_widths(compact, expected_widths, warnings)

    # ─── Strategy C: maybe missing positions; try slicing the compact form ───
    compact = re.sub(r"[\s\-/.]", "", s)
    return _slice_by_widths(compact, expected_widths, warnings)


def _slice_by_widths(s: str, widths: list[int], warnings: list[str]) -> tuple[list[str], list[str]]:
    """Slice a compact string into tokens of given widths."""
    tokens = []
    pos = 0
    for w in widths:
        if pos + w > len(s):
            tokens.append(s[pos:] if pos < len(s) else "")
            pos = len(s)
        else:
            tokens.append(s[pos:pos + w])
            pos += w
    if pos < len(s):
        warnings.append(f"Trailing characters not assigned to any position: '{s[pos:]}'")
    return tokens, warnings


# ─────────────────────────────────────────────────────────────────────────────
# Lookup resolution
# ─────────────────────────────────────────────────────────────────────────────
def _resolve_field(layout: SheetLayout, pg: PositionGroup, code: str) -> dict:
    """
    Look up a single decoded field.
    Returns a dict with: positions, group, code, description, plus any extra columns
    (e.g. parsed displacement_cc).
    """
    out = {
        "positions": pg.positions,
        "group": pg.group,
        "code": code,
        "description": "",
        "recognised": False,
    }

    section_title = pg.lookup_section

    # Fixed positions (product, series identifier, etc.) don't have a lookup section
    if section_title.startswith("(fixed"):
        # Mark recognised if it matches the example in the template
        if pg.example and code.upper() == pg.example.upper():
            out["recognised"] = True
            out["description"] = pg.defines
        else:
            out["description"] = f"(fixed group: expected '{pg.example}', got '{code}')"
        return out

    # Find the section in lookups
    section = layout.lookups.get(section_title)
    if section is None:
        # Try a looser match (some users may have edited section titles)
        for title in layout.lookups:
            if title.lower().startswith(section_title.lower()[:12]):
                section = layout.lookups[title]
                break

    if section is None:
        out["description"] = f"(lookup section '{section_title}' not found)"
        return out

    row = section.get(code)
    if row is None:
        out["description"] = f"(code '{code}' not found in {pg.lookup_section})"
        return out

    out["recognised"] = True
    # Description is the first non-key column we find
    for k, v in row.items():
        if v is None:
            continue
        # Skip the 'Description' column header that just duplicates the description we set
        if str(k).strip().lower() == "description":
            if not out["description"] and isinstance(v, str):
                out["description"] = v
            continue
        # Carry every column through (so e.g. Displacement in cm³ and in³ are both kept)
        out[k] = v
        if not out["description"] and isinstance(v, str):
            out["description"] = v

    return out


# ─────────────────────────────────────────────────────────────────────────────
# Spec normalisation — convert decoded fields into spec_matching.py keys
# ─────────────────────────────────────────────────────────────────────────────
def _normalise_specs(decoded_fields: list[dict], series: str) -> tuple[dict, list[str]]:
    """
    Turn the decoded fields into a flat spec dict using the same key names
    that spec_matching.find_matches() expects.

    Returns (specs_dict, warnings).
    """
    specs = {
        "displacement_cc": None,
        "max_pressure_bar": None,
        "max_speed_rpm": None,
        "max_flow_lpm": None,
        "max_torque_nm": None,
        "shaft_pref": None,
        "shaft_diameter_mm": None,
        "mount_pref": None,
        "port_pref": None,
        "rotation": "CW",
        "special": None,
    }
    notes = []

    fields_by_group = {f["group"]: f for f in decoded_fields}

    # ─── Displacement ───
    disp_field = fields_by_group.get("Displacement")
    if disp_field and disp_field.get("recognised"):
        # Try both possible column names from the two sheets
        cc = (disp_field.get("Displacement (cm³/rev)")
              or disp_field.get("Displacement (cm3/rev)"))
        if cc is not None:
            specs["displacement_cc"] = float(cc)

    # ─── Mount type — infer category from description ───
    mount_field = fields_by_group.get("Mounting type") or fields_by_group.get("Mount")
    if mount_field and mount_field.get("recognised"):
        desc = (mount_field.get("description") or "").lower()
        if "wheel" in desc:
            specs["mount_pref"] = "wheel"
        elif "magneto" in desc:
            specs["mount_pref"] = "magneto"
        elif "bearingless" in desc:
            specs["mount_pref"] = "bearingless"
        elif "sae a" in desc or "sae type a" in desc:
            specs["mount_pref"] = "SAE A"
        elif "sae b" in desc:
            specs["mount_pref"] = "SAE B"
        elif "manifold" in desc:
            specs["mount_pref"] = "manifold"
        elif re.search(r"\b4 bolt\b", desc) and "82.6" in desc:
            # 4-bolt square pattern around SAE A pilot
            specs["mount_pref"] = "SAE A"
        elif re.search(r"\b2 bolt\b", desc):
            specs["mount_pref"] = "SAE A"
        else:
            notes.append(f"Mount type could not be cleanly classified from: '{desc[:60]}…'")

    # ─── Shaft — parse type + diameter ───
    shaft_field = (fields_by_group.get("Output shaft")
                   or fields_by_group.get("Output Shaft"))
    if shaft_field and shaft_field.get("recognised"):
        desc = (shaft_field.get("description") or "").lower()
        # Diameter — prefer explicit mm value, e.g. "1 inch (25.4 mm)" → 25.4
        mm_match = re.search(r"\((\d+(?:\.\d+)?)\s*mm\)", desc)
        if mm_match:
            specs["shaft_diameter_mm"] = float(mm_match.group(1))
        else:
            m = re.search(r"(\d+(?:\.\d+)?)\s*\(", desc)
            if m:
                specs["shaft_diameter_mm"] = float(m.group(1))
        # Type
        if "tapered" in desc:
            specs["shaft_pref"] = "tapered"
        elif "spline" in desc:
            specs["shaft_pref"] = "spline"
        elif "straight" in desc and ("key" in desc or "woodruff" in desc):
            specs["shaft_pref"] = "parallel key"
        elif "straight" in desc:
            specs["shaft_pref"] = "straight"
        elif "crosshole" in desc:
            specs["shaft_pref"] = "straight"

    # ─── Port — pull a clean thread spec from the description ───
    port_field = (fields_by_group.get("Ports")
                  or fields_by_group.get("Port type"))
    if port_field and port_field.get("recognised"):
        desc = port_field.get("description") or ""
        port = _extract_port_thread(desc)
        if port:
            specs["port_pref"] = port

    # ─── Rotation — Char-Lynn 2000 has "Reverse rotation" in Special features (assembly) ───
    sf_asm = fields_by_group.get("Special features (asm)") or \
             fields_by_group.get("Special features (assembly)")
    if sf_asm and sf_asm.get("recognised"):
        desc = (sf_asm.get("description") or "").lower()
        if "reverse rotation" in desc:
            specs["rotation"] = "CCW"

    # T Series — "Special assembly instructions" position
    sa = fields_by_group.get("Special assembly instructions")
    if sa and sa.get("recognised"):
        desc = (sa.get("description") or "").lower()
        if "reverse rotation" in desc:
            specs["rotation"] = "CCW"

    # ─── Pressure — infer from Pressure/flow option (2000 Series only) ───
    pf = fields_by_group.get("Pressure / flow option") or \
         fields_by_group.get("Pressure/flow option")
    if pf and pf.get("recognised"):
        desc = pf.get("description") or ""
        # e.g. "Set at 189.6 bar (2750 lbf/in²)"
        m = re.search(r"([\d.]+)\s*bar", desc)
        if m:
            specs["max_pressure_bar"] = float(m.group(1))
            notes.append(
                f"max_pressure_bar = {specs['max_pressure_bar']} inferred from integral "
                f"cross-over relief valve setting (position {pf.get('positions')}). "
                f"This is the relief protection level, not necessarily the motor's continuous rating."
            )

    # If pressure not yet set, leave None and flag it
    if specs["max_pressure_bar"] is None:
        notes.append("max_pressure_bar not encoded in model code — supply from catalog rating table.")

    if specs["max_speed_rpm"] is None:
        notes.append("max_speed_rpm not encoded in model code — supply from catalog rating table.")
    if specs["max_flow_lpm"] is None:
        notes.append("max_flow_lpm not encoded in model code — supply from catalog rating table.")
    if specs["max_torque_nm"] is None:
        notes.append("max_torque_nm not encoded in model code — supply from catalog rating table.")

    return specs, notes


def _extract_port_thread(desc: str) -> Optional[str]:
    """Pull a clean port thread spec out of a description string."""
    d = desc.lower()
    # SAE / UN family — capture optional leading dot/digit
    # e.g. ".875-14 UNF-2B" or "1.0625-12 UN-2B" or "1 1/16-12UNF"
    m = re.search(r"(\.?\d+(?:\.\d+)?(?:\s*\d+/\d+)?-\d+\s*un[ef]?)", d, re.IGNORECASE)
    if m:
        thread = m.group(1).upper().replace(" ", "")
        # Convert common decimal fractions to fraction form (e.g. .875 → 7/8)
        decimal_to_fraction = {
            "0.5": "1/2", ".5": "1/2",
            "0.625": "5/8", ".625": "5/8",
            "0.75": "3/4", ".75": "3/4",
            "0.875": "7/8", ".875": "7/8",
            "0.4375": "7/16", ".4375": "7/16",
            "0.3125": "5/16", ".3125": "5/16",
            "1.0625": "1-1/16",
        }
        for dec, frac in decimal_to_fraction.items():
            if thread.startswith(dec.upper()) or thread.lower().startswith(dec):
                thread = frac + thread[len(dec):]
                break
        return thread
    # G/BSP
    m = re.search(r"\bg[\s-]?(\d+/\d+)", d)
    if m:
        return f"G{m.group(1)}"
    # NPTF
    m = re.search(r"(\.?\d+(?:\.\d+)?-\d+\s*nptf)", d, re.IGNORECASE)
    if m:
        return m.group(1).upper().replace(" ", "")
    # Manifold ports — no thread to extract
    if "manifold" in d:
        return "manifold"
    return None


# ─────────────────────────────────────────────────────────────────────────────
# Result container
# ─────────────────────────────────────────────────────────────────────────────
@dataclass
class CompetitorDecodeResult:
    raw_input: str
    normalised: str
    is_valid: bool
    brand: Optional[str]
    series: Optional[str]
    sheet_name: Optional[str]
    fields: list[dict] = field(default_factory=list)
    specs_for_crossover: dict = field(default_factory=dict)
    unknown_codes: list[str] = field(default_factory=list)
    warnings: list[str] = field(default_factory=list)


# ─────────────────────────────────────────────────────────────────────────────
# Main entry point
# ─────────────────────────────────────────────────────────────────────────────
def decode_competitor_code(raw: str,
                           workbook_path: Optional[str] = None
                           ) -> CompetitorDecodeResult:
    """
    Decode a competitor's orbital motor model code.

    Args:
        raw: the model code string (any common separator style)
        workbook_path: optional path to competitor_code_extractor.xlsx;
                       defaults to the script's folder.

    Returns:
        CompetitorDecodeResult with full decoded fields and a normalised
        spec dict ready for spec_matching.find_matches().
    """
    result = CompetitorDecodeResult(
        raw_input=raw, normalised="", is_valid=False,
        brand=None, series=None, sheet_name=None,
    )

    if not raw or not raw.strip():
        result.warnings.append("Empty input.")
        return result

    # Resolve workbook
    if workbook_path is None:
        workbook_path = str(Path(__file__).parent / "competitor_code_extractor.xlsx")
    if not Path(workbook_path).exists():
        result.warnings.append(f"Workbook not found: {workbook_path}")
        return result

    # Detect series
    detected = _detect_series(raw)
    if not detected:
        result.warnings.append(
            f"Could not identify competitor series. Expected prefixes: "
            f"{[p[0] for p in SERIES_PREFIXES]}. "
            f"Got: {raw.strip()[:8]!r}"
        )
        return result

    prefix, brand, series, sheet_name = detected
    result.brand = brand
    result.series = series
    result.sheet_name = sheet_name

    # Load and parse the matching sheet
    wb = load_workbook(workbook_path, data_only=True)
    if sheet_name not in wb.sheetnames:
        result.warnings.append(f"Sheet '{sheet_name}' missing from workbook.")
        return result
    layout = _parse_sheet(wb, sheet_name, brand, series)

    # Tokenise input
    tokens, tk_warnings = _tokenise(raw, layout)
    result.warnings.extend(tk_warnings)

    if len(tokens) < len(layout.position_groups):
        result.warnings.append(
            f"Input has {len(tokens)} positions; "
            f"{series} expects {len(layout.position_groups)}. "
            f"Missing trailing positions will be marked unrecognised."
        )

    # Resolve each field
    for i, pg in enumerate(layout.position_groups):
        if i < len(tokens):
            code = tokens[i]
        else:
            code = ""
        field_data = _resolve_field(layout, pg, code)
        result.fields.append(field_data)
        if not field_data.get("recognised") and code:
            result.unknown_codes.append(f"{pg.group} (pos {pg.positions}): '{code}'")

    # Build the normalised part-number string
    parts = []
    for i, pg in enumerate(layout.position_groups):
        if i < len(tokens):
            parts.append(tokens[i])
    result.normalised = " ".join(parts)

    # Normalise to crossover spec dict
    specs, spec_notes = _normalise_specs(result.fields, series)
    result.specs_for_crossover = specs
    result.warnings.extend(spec_notes)

    result.is_valid = len([f for f in result.fields if f.get("recognised")]) > 0

    return result


# ─────────────────────────────────────────────────────────────────────────────
# Pretty-printed report
# ─────────────────────────────────────────────────────────────────────────────
def _fmt(v, suffix: str = "") -> str:
    if v is None:
        return "—"
    if isinstance(v, float):
        return f"{v:g}{suffix}"
    return f"{v}{suffix}"


def print_decoded(result: CompetitorDecodeResult) -> None:
    """Pretty-print a decoded competitor model code."""
    print()
    print("═" * 78)
    print(" COMPETITOR MODEL CODE DECODER")
    print("═" * 78)
    print(f"\n  Input:      {result.raw_input}")
    print(f"  Normalised: {result.normalised or '(could not parse)'}")
    print(f"  Brand:      {result.brand or '(unknown)'}")
    print(f"  Series:     {result.series or '(unknown)'}")

    if not result.is_valid:
        print("\n  ❌ Could not decode this model code.")
        for w in result.warnings:
            print(f"     {w}")
        print()
        return

    # ─── Decoded fields ───
    print()
    print("─" * 78)
    print(" DECODED FIELDS")
    print("─" * 78)
    for f in result.fields:
        marker = "  " if f.get("recognised") else "⚠ "
        positions = f.get("positions", "?")
        group = f.get("group", "?")
        code = f.get("code", "") or "(missing)"
        desc = f.get("description", "")
        # Truncate description to ~60 chars on first line, wrap rest
        if len(desc) > 60:
            print(f"  {marker}POS {positions:<6} {group:<30} {code:<5}  {desc[:60]}…")
        else:
            print(f"  {marker}POS {positions:<6} {group:<30} {code:<5}  {desc}")
        # Show extra columns if present (e.g. Displacement cm³/in³)
        extras = {k: v for k, v in f.items()
                  if k not in ("positions", "group", "code", "description",
                               "recognised", "Description") and v is not None}
        if extras and len(extras) <= 4:
            for k, v in extras.items():
                if isinstance(v, str) and len(v) > 80:
                    continue   # skip the long description duplicate
                print(f"                                              {k}: {v}")

    # ─── Specs for crossover ───
    print()
    print("─" * 78)
    print(" NORMALISED SPECS (for spec_matching.find_matches)")
    print("─" * 78)
    specs = result.specs_for_crossover
    key_order = [
        "displacement_cc", "max_pressure_bar", "max_speed_rpm",
        "max_flow_lpm", "max_torque_nm",
        "mount_pref", "shaft_pref", "shaft_diameter_mm",
        "port_pref", "rotation", "special",
    ]
    for k in key_order:
        v = specs.get(k)
        marker = "  " if v not in (None, "") else "○ "
        print(f"  {marker}{k:<22} {_fmt(v)}")

    # ─── Unknown codes ───
    if result.unknown_codes:
        print()
        print("─" * 78)
        print(" UNRECOGNISED CODES")
        print("─" * 78)
        for uk in result.unknown_codes:
            print(f"  ⚠ {uk}")

    # ─── Warnings ───
    if result.warnings:
        print()
        print("─" * 78)
        print(" NOTES / WARNINGS")
        print("─" * 78)
        for w in result.warnings:
            print(f"  • {w}")

    print()
    print("═" * 78)
    if not result.unknown_codes:
        print(" ✅ Model code decoded successfully.")
    else:
        print(" ⚠  Model code decoded with unrecognised codes.")
    print("═" * 78)
    print()


# ─────────────────────────────────────────────────────────────────────────────
# CLI
# ─────────────────────────────────────────────────────────────────────────────
if __name__ == "__main__":
    xlsx_path = Path(__file__).parent / "competitor_code_extractor.xlsx"
    if not xlsx_path.exists():
        print(f"❌ Could not find {xlsx_path}")
        print("   Make sure competitor_code_extractor.xlsx is in the same folder.")
        sys.exit(1)

    if len(sys.argv) > 1:
        code = " ".join(sys.argv[1:])
        result = decode_competitor_code(code, str(xlsx_path))
        print_decoded(result)
    else:
        print("Competitor Model Code Decoder")
        print("(enter blank or 'quit' to exit)")
        while True:
            try:
                code = input("\nModel code: ").strip()
            except (EOFError, KeyboardInterrupt):
                print()
                break
            if not code or code.lower() in ("quit", "exit", "q"):
                break
            result = decode_competitor_code(code, str(xlsx_path))
            print_decoded(result)
