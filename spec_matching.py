"""
Hengli Orbital Motor Crossover Tool
====================================

Given a set of specs (typically from a competitor's catalog), find the best-matching
Hengli orbital motor model(s) and generate a realistic part number.

USAGE
-----
    from hengli_crossover import MotorDatabase, find_matches, print_comparison

    db = MotorDatabase("Hengli_Orbital_Motor_Master.xlsx")

    specs = {
        "displacement_cc": 100,        # required (or flow+speed)
        "max_pressure_bar": 200,       # strongly recommended
        "max_speed_rpm": 600,          # optional
        "max_flow_lpm": 75,            # optional
        "max_torque_nm": 300,          # optional
        "shaft_pref": "spline",        # optional: 'spline', 'parallel key', 'tapered'
        "shaft_diameter_mm": 25.4,     # optional
        "mount_pref": "SAE A",         # optional: 'SAE A', 'SAE B', 'magneto', 'square', 'wheel'
        "port_pref": "G1/2",           # optional: thread spec or rough form
        "rotation": "CW",              # optional: 'CW' or 'CCW'
    }

    matches = find_matches(specs, db, top_n=3)
    print_comparison(matches, specs)

REQUIREMENTS
------------
- Python 3.9+
- openpyxl

The Excel workbook must follow the layout produced by the master-sheet build
(sheets: Models, Mount_Port_Codes, Port_Codes, Shaft_Codes, Option_Codes, Notes).
"""

from __future__ import annotations
from dataclasses import dataclass, field
from typing import Optional
from openpyxl import load_workbook
import re


# ─────────────────────────────────────────────────────────────────────────────
# Column map for the Models sheet (row 2/3 header, data starts row 4)
# ─────────────────────────────────────────────────────────────────────────────
MODELS_COLS = {
    "brand": 1, "series": 2, "type": 3, "distribution": 4,
    "displacement_cc": 5,
    "max_speed_cont": 6, "max_speed_inter": 7,
    "max_torque_cont": 8, "max_torque_inter": 9,
    "max_output_cont": 10, "max_output_inter": 11,
    "max_dp_cont": 12, "max_dp_inter": 13, "max_dp_peak": 14,
    "max_flow_cont": 15, "max_flow_inter": 16,
    "max_no_load_start_p": 17,
    "min_start_torque_cont": 18, "min_start_torque_inter": 19,
    "weight_standard": 20, "weight_bearingless": 21,
    "mount_code_field": 22, "port_code_field": 23, "shaft_code_field": 24,
    "rotation_code_field": 25, "paint_code_field": 26, "special_code_field": 27,
    "example_pn": 28,
}


# ═════════════════════════════════════════════════════════════════════════════
# Database loader
# ═════════════════════════════════════════════════════════════════════════════
class MotorDatabase:
    """Loads the Hengli master workbook into flat Python structures."""

    def __init__(self, xlsx_path: str):
        self.xlsx_path = xlsx_path
        self.models: list[dict] = []
        self.mounts: list[dict] = []
        self.ports: list[dict] = []
        self.shafts: list[dict] = []
        self.options: list[dict] = []
        self._load()

    def _load(self):
        wb = load_workbook(self.xlsx_path, data_only=True)

        # ----- Models -----
        ws = wb["Models"]
        for r in range(4, ws.max_row + 1):
            brand = ws.cell(row=r, column=1).value
            if not brand:
                continue
            model = {k: ws.cell(row=r, column=c).value for k, c in MODELS_COLS.items()}
            # Normalise type to string ("050" not 50)
            model["type"] = str(model["type"])
            self.models.append(model)

        # ----- Mount_Port_Codes -----
        ws = wb["Mount_Port_Codes"]
        for r in range(3, ws.max_row + 1):
            series = ws.cell(row=r, column=1).value
            if not series:
                continue
            self.mounts.append({
                "series": series,
                "code": ws.cell(row=r, column=2).value,
                "description": ws.cell(row=r, column=3).value or "",
                "port": ws.cell(row=r, column=4).value or "",
                "drain": ws.cell(row=r, column=5).value or "",
            })

        # ----- Port_Codes (HSD/HSP/HSE only — HRD has port combined with mount) -----
        ws = wb["Port_Codes"]
        for r in range(3, ws.max_row + 1):
            series = ws.cell(row=r, column=1).value
            if not series:
                continue
            self.ports.append({
                "series": series,
                "code": ws.cell(row=r, column=2).value,
                "port_ab": ws.cell(row=r, column=3).value or "",
                "drain": ws.cell(row=r, column=4).value or "",
                "notes": ws.cell(row=r, column=5).value or "",
            })

        # ----- Shaft_Codes -----
        ws = wb["Shaft_Codes"]
        for r in range(3, ws.max_row + 1):
            series = ws.cell(row=r, column=1).value
            if not series:
                continue
            self.shafts.append({
                "series": series,
                "code": ws.cell(row=r, column=2).value,
                "description": ws.cell(row=r, column=3).value or "",
                "max_torque_nm": ws.cell(row=r, column=4).value,
                "compatible_mounts": ws.cell(row=r, column=5).value or "",
            })

        # ----- Option_Codes -----
        ws = wb["Option_Codes"]
        for r in range(3, ws.max_row + 1):
            series = ws.cell(row=r, column=1).value
            if not series:
                continue
            self.options.append({
                "series": series,
                "position": ws.cell(row=r, column=2).value,
                "code": ws.cell(row=r, column=3).value,
                "description": ws.cell(row=r, column=4).value or "",
                "notes": ws.cell(row=r, column=5).value or "",
            })

    def mounts_for_series(self, series: str) -> list[dict]:
        return [m for m in self.mounts if m["series"] == series]

    def ports_for_series(self, series: str) -> list[dict]:
        return [p for p in self.ports if p["series"] == series]

    def shafts_for_series(self, series: str) -> list[dict]:
        return [s for s in self.shafts if s["series"] == series]

    def options_for_series(self, series: str, position: str = None) -> list[dict]:
        return [o for o in self.options
                if o["series"] == series
                and (position is None or o["position"] == position)]


# ═════════════════════════════════════════════════════════════════════════════
# Input sufficiency check
# ═════════════════════════════════════════════════════════════════════════════
def check_input_sufficiency(specs: dict) -> tuple[bool, str]:
    """
    Verify the user gave us enough to make a sensible match.

    Hard requirement: either displacement_cc, or both max_flow_lpm and max_speed_rpm.

    Returns (is_sufficient, message).
    """
    have_disp = specs.get("displacement_cc") is not None
    have_flow = specs.get("max_flow_lpm") is not None
    have_speed = specs.get("max_speed_rpm") is not None

    if have_disp:
        return True, f"Displacement provided ({specs['displacement_cc']} cc/rev)."

    if have_flow and have_speed:
        # Derive a working displacement estimate
        disp_est = specs["max_flow_lpm"] * 1000.0 / specs["max_speed_rpm"]
        return True, (f"Displacement not given but derivable from flow and speed: "
                      f"~{disp_est:.1f} cc/rev (= {specs['max_flow_lpm']} L/min × 1000 ÷ "
                      f"{specs['max_speed_rpm']} rpm).")

    missing = []
    if not have_disp:
        missing.append("displacement_cc")
    if not (have_flow and have_speed):
        missing.append("max_flow_lpm AND max_speed_rpm (as fallback for displacement)")

    return False, (
        "Not enough information to match. Need at least one of:\n"
        "  • displacement_cc (cm³/rev), OR\n"
        "  • max_flow_lpm + max_speed_rpm (so displacement can be derived).\n"
        f"Missing: {', '.join(missing)}\n\n"
        "Also strongly recommended: max_pressure_bar (without it, fit-rating is unreliable)."
    )


# ═════════════════════════════════════════════════════════════════════════════
# Match scoring
# ═════════════════════════════════════════════════════════════════════════════
@dataclass
class MatchResult:
    model: dict                      # raw model row from the database
    score: float                     # 0–100, higher is better
    fit_rating: str                  # 'Direct', 'Functional', 'Approximate'
    notes: list[str] = field(default_factory=list)
    caveats: list[str] = field(default_factory=list)
    suggested_part_number: str = ""
    chosen_codes: dict = field(default_factory=dict)  # mount/port/shaft picks


def _resolve_displacement(specs: dict) -> float:
    if specs.get("displacement_cc") is not None:
        return float(specs["displacement_cc"])
    return specs["max_flow_lpm"] * 1000.0 / specs["max_speed_rpm"]


def _score_model(model: dict, specs: dict, target_disp: float) -> tuple[float, list[str], list[str]]:
    """
    Score 0–100 for how well this model matches the request.

    Hard cut: pressure rating must meet the request (otherwise score = 0).
    Soft scoring: displacement closeness, speed/flow/torque headroom, weight, etc.
    """
    notes: list[str] = []
    caveats: list[str] = []

    # ----- Hard requirement: pressure -----
    user_pressure = specs.get("max_pressure_bar")
    if user_pressure is not None:
        # Use the model's continuous Δp as the headline rating.
        model_pressure = model.get("max_dp_cont")
        if model_pressure is None:
            caveats.append("Model has no continuous Δp listed — cannot verify pressure rating.")
        elif model_pressure < user_pressure:
            # Hard fail: insufficient pressure
            return 0.0, [], [f"Insufficient pressure rating: model {model_pressure} bar < required {user_pressure} bar."]

    # ----- Displacement closeness (most important soft score) -----
    model_disp = model.get("displacement_cc")
    if model_disp is None:
        return 0.0, [], ["Model has no displacement listed."]

    disp_ratio = model_disp / target_disp
    disp_diff_pct = abs(model_disp - target_disp) / target_disp * 100

    if disp_diff_pct <= 5:
        disp_score = 50.0
    elif disp_diff_pct <= 15:
        disp_score = 50.0 - (disp_diff_pct - 5) * 2.0     # 30..50
    elif disp_diff_pct <= 30:
        disp_score = 30.0 - (disp_diff_pct - 15) * 1.5    # 7.5..30
    else:
        disp_score = max(0.0, 7.5 - (disp_diff_pct - 30) * 0.3)

    if disp_diff_pct > 10:
        caveats.append(
            f"Displacement differs by {disp_diff_pct:.0f}% "
            f"(model {model_disp:g} vs requested {target_disp:.1f} cc/rev)."
        )

    # ----- Speed headroom (continuous rpm should meet or exceed need) -----
    speed_score = 0.0
    user_speed = specs.get("max_speed_rpm")
    if user_speed is not None and model.get("max_speed_cont"):
        if model["max_speed_cont"] >= user_speed:
            # Reward closeness without penalising headroom too much
            ratio = user_speed / model["max_speed_cont"]
            speed_score = 10.0 * min(1.0, ratio + 0.5)   # caps at 10
        else:
            speed_score = 0.0
            caveats.append(
                f"Speed shortfall: continuous {model['max_speed_cont']} rpm < requested {user_speed} rpm."
            )

    # ----- Flow headroom -----
    flow_score = 0.0
    user_flow = specs.get("max_flow_lpm")
    if user_flow is not None and model.get("max_flow_cont"):
        if model["max_flow_cont"] >= user_flow:
            flow_score = 10.0
        elif model.get("max_flow_inter") and model["max_flow_inter"] >= user_flow:
            flow_score = 5.0
            caveats.append(
                f"Flow {user_flow} L/min exceeds continuous rating ({model['max_flow_cont']}), "
                f"within intermittent ({model['max_flow_inter']}). Limit duty cycle to <6s/min."
            )
        else:
            flow_score = 0.0
            caveats.append(
                f"Flow shortfall: continuous {model['max_flow_cont']} L/min < requested {user_flow} L/min."
            )

    # ----- Torque headroom -----
    torque_score = 0.0
    user_torque = specs.get("max_torque_nm")
    if user_torque is not None and model.get("max_torque_cont"):
        if model["max_torque_cont"] >= user_torque:
            torque_score = 10.0
        elif model.get("max_torque_inter") and model["max_torque_inter"] >= user_torque:
            torque_score = 5.0
            caveats.append(
                f"Torque {user_torque} N·m exceeds continuous rating ({model['max_torque_cont']}), "
                f"within intermittent ({model['max_torque_inter']}). Limit duty cycle."
            )
        else:
            caveats.append(
                f"Torque shortfall: continuous {model['max_torque_cont']} N·m < requested {user_torque} N·m."
            )

    # ----- Pressure headroom bonus (model has more pressure than asked) -----
    pressure_bonus = 0.0
    if user_pressure is not None and model.get("max_dp_cont"):
        if model["max_dp_cont"] >= user_pressure:
            # Reward slight headroom but don't reward gross overspec
            headroom = (model["max_dp_cont"] - user_pressure) / user_pressure
            if headroom <= 0.30:
                pressure_bonus = 10.0
            elif headroom <= 0.60:
                pressure_bonus = 7.0
            else:
                pressure_bonus = 4.0
                notes.append(
                    f"Pressure overspec ({model['max_dp_cont']} bar vs {user_pressure} bar required) — "
                    "consider a lighter/cheaper series."
                )

    # ----- Mount preference bonus -----
    mount_bonus = 0.0
    user_mount = specs.get("mount_pref")
    if user_mount:
        # We don't pin a mount code yet, just check if the series has a matching mount
        # (this will be refined when we pick the actual code).
        mount_bonus = 5.0  # speculative; will be confirmed when codes are chosen

    # ----- Shaft preference bonus (same: speculative pending code selection) -----
    shaft_bonus = 0.0
    user_shaft = specs.get("shaft_pref") or specs.get("shaft_diameter_mm")
    if user_shaft:
        shaft_bonus = 5.0

    total = disp_score + speed_score + flow_score + torque_score + pressure_bonus + mount_bonus + shaft_bonus
    total = max(0.0, min(100.0, total))

    # Add positive notes
    if model.get("distribution"):
        notes.append(f"Distribution: {model['distribution']}.")

    return total, notes, caveats


def _classify_fit(score: float, caveats: list[str]) -> str:
    hard_problems = any(
        "shortfall" in c.lower() or "insufficient" in c.lower()
        for c in caveats
    )
    if hard_problems:
        return "Approximate"
    if score >= 80:
        return "Direct"
    if score >= 55:
        return "Functional"
    return "Approximate"


# ═════════════════════════════════════════════════════════════════════════════
# Code selection (mount, port, shaft) — picks a realistic part number
# ═════════════════════════════════════════════════════════════════════════════
def _normalise(s: str) -> str:
    return (s or "").lower()


def _pick_mount(series: str, db: MotorDatabase, user_mount: Optional[str]) -> Optional[dict]:
    candidates = db.mounts_for_series(series)
    if not candidates:
        return None

    if user_mount:
        u = _normalise(user_mount)
        # Try keyword match against the description
        keyword_map = {
            "sae a": ["sae a", "sae type a"],
            "sae b": ["sae b", "sae type b"],
            "magneto": ["magneto"],
            "square": ["square"],
            "wheel": ["wheel"],
            "round": ["round"],
        }
        for key, kws in keyword_map.items():
            if any(k in u for k in kws):
                for m in candidates:
                    desc = _normalise(m["description"])
                    if any(k in desc for k in kws):
                        return m

    # Defaults per series (most common general-purpose mount)
    series_defaults = {
        "HRD": "A23",   # SAE A 2-hole, standard
        "HSD": "A3",
        "HSP": "A3",
        "HSE": "F1",
    }
    default_code = series_defaults.get(series)
    for m in candidates:
        if m["code"] == default_code:
            return m

    return candidates[0]


def _pick_port(series: str, db: MotorDatabase, mount: Optional[dict], user_port: Optional[str]) -> Optional[dict]:
    # HRD encodes the port inside the mount code — no separate port code
    if series == "HRD":
        return None

    candidates = db.ports_for_series(series)
    if not candidates:
        return None

    if user_port:
        u = _normalise(user_port)
        # Try direct thread-spec match
        for p in candidates:
            if u in _normalise(p["port_ab"]):
                return p
        # Try fuzzy thread-family
        if "g1/2" in u or "bsp" in u:
            for p in candidates:
                if "g1/2" in _normalise(p["port_ab"]):
                    return p
        if "unf" in u or "7/8" in u:
            for p in candidates:
                if "7/8-14unf" in _normalise(p["port_ab"]):
                    return p
        if "m22" in u or "metric" in u:
            for p in candidates:
                if "m22" in _normalise(p["port_ab"]):
                    return p

    # Defaults: prefer G1/2 if available, else 7/8-14UNF, else first
    for preferred in ["G1/2", "7/8-14UNF"]:
        for p in candidates:
            if preferred.lower() in _normalise(p["port_ab"]):
                return p

    return candidates[0]


def _pick_shaft(series: str, db: MotorDatabase, mount_code: Optional[str],
                user_shaft_pref: Optional[str], user_shaft_dia: Optional[float]) -> Optional[dict]:
    candidates = db.shafts_for_series(series)
    if not candidates:
        return None

    # Filter by mount compatibility
    if mount_code:
        compatible = []
        for s in candidates:
            cm = s["compatible_mounts"] or ""
            # The compatible_mounts field uses formats like "A23/A24/..." or "Wheel mounts W1/W4 only" etc.
            # A loose check: if the mount code appears in the string, or "All" appears.
            cm_l = cm.lower()
            if "all" in cm_l and "except" not in cm_l:
                compatible.append(s)
            elif "except" in cm_l:
                # "All HSP mounts EXCEPT W1/W4" — include if mount NOT in the exclusion list
                if mount_code.lower() not in cm_l.split("except")[1]:
                    compatible.append(s)
            elif mount_code in cm:
                compatible.append(s)
            elif mount_code.lower() in cm_l:
                compatible.append(s)
        if compatible:
            candidates = compatible

    # Filter by shaft type preference
    if user_shaft_pref:
        u = _normalise(user_shaft_pref)
        type_kws = {
            "spline": ["spline"],
            "key": ["parallel key", "key"],
            "parallel": ["parallel key"],
            "straight": ["straight"],
            "tapered": ["tapered"],
            "taper": ["tapered"],
            "cardan": ["cardan"],
        }
        for typ, kws in type_kws.items():
            if typ in u:
                filtered = [s for s in candidates
                            if any(k in _normalise(s["description"]) for k in kws)]
                if filtered:
                    candidates = filtered
                    break

    # Filter by shaft diameter — prefer exact match, fall back to ±1.0 mm only if nothing exact
    if user_shaft_dia:
        exact_match = []
        loose_match = []
        for s in candidates:
            m = re.search(r"(?:ø\s*)?([0-9]+(?:\.[0-9]+)?)\s*(?:mm|\(|straight|spline|taper)", _normalise(s["description"]))
            if m:
                dia = float(m.group(1))
                if abs(dia - user_shaft_dia) <= 0.1:
                    exact_match.append(s)
                elif abs(dia - user_shaft_dia) <= 1.0:
                    loose_match.append(s)
        if exact_match:
            candidates = exact_match
        elif loose_match:
            candidates = loose_match

    if not candidates:
        return None

    # Default preference: the first one (these are typically listed in popularity order)
    # Series-specific sensible default:
    series_default_shafts = {
        "HRD": "S2",
        "HSD": "S1",
        "HSP": "S3",
        "HSE": "R1",
    }
    default_code = series_default_shafts.get(series)
    if default_code:
        for s in candidates:
            if s["code"] == default_code:
                return s

    return candidates[0]


def _build_part_number(model: dict, mount: Optional[dict], port: Optional[dict],
                       shaft: Optional[dict], rotation: str, paint: str,
                       special: str) -> str:
    series = model["series"]
    typ = model["type"]
    rot_code = "A" if rotation.upper() in ("CW", "A") else "R"
    paint_code = paint
    special_code = special

    if series == "HRD":
        # 7 positions: HRD-TYPE-MOUNT-SHAFT-ROT-PAINT-SPECIAL
        parts = ["HRD", typ,
                 mount["code"] if mount else "A23",
                 shaft["code"] if shaft else "S2",
                 rot_code, paint_code, special_code]
    else:
        # 8 positions: SERIES-TYPE-MOUNT-PORT-SHAFT-ROT-PAINT-SPECIAL
        parts = [series, typ,
                 mount["code"] if mount else "?",
                 port["code"] if port else "?",
                 shaft["code"] if shaft else "?",
                 rot_code, paint_code, special_code]
    return "-".join(str(p) for p in parts)


# ═════════════════════════════════════════════════════════════════════════════
# Main entry point
# ═════════════════════════════════════════════════════════════════════════════
def find_matches(specs: dict, db: MotorDatabase, top_n: int = 3,
                 series=None) -> list[MatchResult]:
    """
    Score every model and return the top N. Each result has full specs, fit
    rating, caveats, and a suggested part number with mount/port/shaft codes.

    Args:
        specs: user spec dict (displacement_cc, max_pressure_bar, etc.)
        db: a loaded MotorDatabase
        top_n: how many candidates to return (sorted best-first)
        series: optional series filter. Pass a single series code (e.g. "HSP")
                or a list/set of codes (e.g. ["HSP", "HSE"]) to restrict the
                search to those Hengli series only. Default None = all series.
    """
    ok, _ = check_input_sufficiency(specs)
    if not ok:
        raise ValueError(check_input_sufficiency(specs)[1])

    target_disp = _resolve_displacement(specs)

    rotation = specs.get("rotation", "CW")
    paint = specs.get("paint", "N")
    special = specs.get("special", "A")

    # Series filter — restrict to mapped Hengli series if caller specified one
    candidate_models = db.models
    if series is not None:
        wanted = {series} if isinstance(series, str) else set(series)
        candidate_models = [m for m in db.models if m["series"] in wanted]

    results = []
    for model in candidate_models:
        score, notes, caveats = _score_model(model, specs, target_disp)
        if score == 0.0:
            continue

        # Pick concrete codes for this model
        mount = _pick_mount(model["series"], db, specs.get("mount_pref"))
        port = _pick_port(model["series"], db, mount, specs.get("port_pref"))
        shaft = _pick_shaft(model["series"], db,
                            mount["code"] if mount else None,
                            specs.get("shaft_pref"),
                            specs.get("shaft_diameter_mm"))

        pn = _build_part_number(model, mount, port, shaft, rotation, paint, special)

        # Verify shaft torque headroom if user gave torque
        if shaft and shaft.get("max_torque_nm") and specs.get("max_torque_nm"):
            if shaft["max_torque_nm"] < specs["max_torque_nm"]:
                caveats.append(
                    f"Chosen shaft {shaft['code']} max torque "
                    f"{shaft['max_torque_nm']} N·m < required {specs['max_torque_nm']} N·m. "
                    f"Consider a stronger shaft option."
                )

        chosen = {
            "mount_code": mount["code"] if mount else None,
            "mount_desc": mount["description"] if mount else None,
            "port_code": port["code"] if port else None,
            "port_desc": port["port_ab"] if port else None,
            "shaft_code": shaft["code"] if shaft else None,
            "shaft_desc": shaft["description"] if shaft else None,
            "shaft_max_torque": shaft.get("max_torque_nm") if shaft else None,
            "rotation_code": "A" if rotation.upper() in ("CW", "A") else "R",
            "paint_code": paint,
            "special_code": special,
        }

        fit = _classify_fit(score, caveats)
        results.append(MatchResult(
            model=model, score=round(score, 1), fit_rating=fit,
            notes=notes, caveats=caveats,
            suggested_part_number=pn, chosen_codes=chosen,
        ))

    results.sort(key=lambda r: r.score, reverse=True)

    # 1. 丢弃分数 < 50 的候选
    MIN_SCORE = 50.0
    results = [r for r in results if r.score >= MIN_SCORE]
    if not results:
        return []

    # 2. 排量锁定 + 多样性:以最高分模型的排量为锚点,±10% 内择优,
    #    每个排量值只保留最高分那一个
    anchor_disp = results[0].model.get("displacement_cc")
    if anchor_disp is None:
        return results[:top_n]

    pool = [r for r in results
            if r.model.get("displacement_cc") is not None
            and abs(r.model["displacement_cc"] - anchor_disp) / anchor_disp <= 0.10]

    seen_disps: set[float] = set()
    diverse: list[MatchResult] = []
    for r in pool:
        d = r.model["displacement_cc"]
        if d not in seen_disps:
            seen_disps.add(d)
            diverse.append(r)
        if len(diverse) == top_n:
            break

    return diverse


# ═════════════════════════════════════════════════════════════════════════════
# Report formatting
# ═════════════════════════════════════════════════════════════════════════════
def _fmt(v, suffix: str = "") -> str:
    if v is None:
        return "—"
    if isinstance(v, float):
        return f"{v:g}{suffix}"
    return f"{v}{suffix}"


def print_comparison(matches: list[MatchResult], specs: dict) -> None:
    """Pretty-printed report — single match summary or side-by-side comparison."""
    if not matches:
        print("\n❌ No matching Hengli orbital motor found.\n")
        print("Likely reasons:")
        print("  • The requested pressure exceeds the highest-rated frame size.")
        print("  • Displacement is far outside the 50–985 cm³/rev range covered by HRD/HSD/HSP/HSE.")
        print("  • The combination of specs creates conflicting requirements.")
        return

    target_disp = _resolve_displacement(specs)

    print()
    print("═" * 78)
    print(" HENGLI ORBITAL MOTOR CROSSOVER REPORT")
    print("═" * 78)
    print("\nINPUT SPECS:")
    for k, v in specs.items():
        print(f"  {k:<22} {v}")
    print(f"  {'→ working displacement':<22} {target_disp:.1f} cc/rev")

    print(f"\nFound {len(matches)} candidate{'s' if len(matches) > 1 else ''}.")

    for i, m in enumerate(matches, 1):
        print()
        print("─" * 78)
        print(f" CANDIDATE #{i}   |   Part No.: {m.suggested_part_number}   |   "
              f"Fit: {m.fit_rating}   |   Score: {m.score}/100")
        print("─" * 78)

        mdl = m.model
        print(f"  Series / Frame:      {mdl['series']} / {mdl['type']}")
        print(f"  Distribution:        {mdl.get('distribution', '—')}")
        print(f"  Displacement:        {_fmt(mdl.get('displacement_cc'))} cc/rev")

        print(f"\n  RATINGS                Continuous   Intermittent   Peak")
        print(f"  Max speed (rpm)        {_fmt(mdl.get('max_speed_cont')):>10}   "
              f"{_fmt(mdl.get('max_speed_inter')):>12}   —")
        print(f"  Max torque (N·m)       {_fmt(mdl.get('max_torque_cont')):>10}   "
              f"{_fmt(mdl.get('max_torque_inter')):>12}   —")
        print(f"  Max Δp (bar)           {_fmt(mdl.get('max_dp_cont')):>10}   "
              f"{_fmt(mdl.get('max_dp_inter')):>12}   "
              f"{_fmt(mdl.get('max_dp_peak')):>5}")
        print(f"  Max flow (L/min)       {_fmt(mdl.get('max_flow_cont')):>10}   "
              f"{_fmt(mdl.get('max_flow_inter')):>12}   —")
        print(f"  Max output (kW)        {_fmt(mdl.get('max_output_cont')):>10}   "
              f"{_fmt(mdl.get('max_output_inter')):>12}   —")
        print(f"\n  Min starting torque (N·m): cont @ max Δp = {_fmt(mdl.get('min_start_torque_cont'))}, "
              f"inter @ max Δp = {_fmt(mdl.get('min_start_torque_inter'))}")
        print(f"  No-load starting pressure: {_fmt(mdl.get('max_no_load_start_p'))} bar")
        print(f"  Weight:               standard {_fmt(mdl.get('weight_standard'))} kg, "
              f"bearingless {_fmt(mdl.get('weight_bearingless'))} kg")

        print(f"\n  CHOSEN CODES (auto-picked; override via the spec dict):")
        cc = m.chosen_codes
        print(f"  Mount    {cc['mount_code']:<6}  {cc['mount_desc']}")
        if cc.get("port_code"):
            print(f"  Port     {cc['port_code']:<6}  {cc['port_desc']}")
        else:
            print(f"  Port     (combined with mount — HRD style)")
        print(f"  Shaft    {cc['shaft_code']:<6}  {cc['shaft_desc']}")
        if cc.get("shaft_max_torque"):
            print(f"           (shaft max torque: {cc['shaft_max_torque']} N·m)")
        print(f"  Rotation {cc['rotation_code']}     Paint {cc['paint_code']}     "
              f"Special {cc['special_code']}")

        if m.notes:
            print("\n  Notes:")
            for n in m.notes:
                print(f"    • {n}")

        if m.caveats:
            print("\n  Caveats:")
            for c in m.caveats:
                print(f"    ⚠ {c}")

    print()
    print("═" * 78)
    if len(matches) > 1:
        print("COMPARISON SUMMARY:")
        print(f"  {'Rank':<6}{'Part Number':<28}{'Disp':<10}{'Δp cont':<10}{'Fit':<14}{'Score'}")
        for i, m in enumerate(matches, 1):
            print(f"  #{i:<5}{m.suggested_part_number:<28}"
                  f"{_fmt(m.model.get('displacement_cc'), ' cc'):<10}"
                  f"{_fmt(m.model.get('max_dp_cont'), ' bar'):<10}"
                  f"{m.fit_rating:<14}{m.score}")
    print("═" * 78)
    print()


# ═════════════════════════════════════════════════════════════════════════════
# CLI demo
# ═════════════════════════════════════════════════════════════════════════════
if __name__ == "__main__":
    import sys

    xlsx_path = sys.argv[1] if len(sys.argv) > 1 else "Hengli_Orbital_Motor_Master.xlsx"

    db = MotorDatabase(xlsx_path)
    print(f"Loaded {len(db.models)} motor models from {xlsx_path}")
    print(f"  Series: {sorted(set(m['series'] for m in db.models))}")
    print(f"  Mount codes: {len(db.mounts)}, Port codes: {len(db.ports)}, "
          f"Shaft codes: {len(db.shafts)}, Option rows: {len(db.options)}")

    # Example query — crossing over a Danfoss OMP 100 (spool valve, ~100 cc, 175 bar)
    example_specs = {
        "displacement_cc": 100,
        "max_pressure_bar": 175,
        "max_speed_rpm": 600,
        "max_flow_lpm": 60,
        "max_torque_nm": 250,
        "shaft_pref": "parallel key",
        "shaft_diameter_mm": 25.4,
        "mount_pref": "SAE A",
        "port_pref": "G1/2",
        "rotation": "CW",
    }

    ok, msg = check_input_sufficiency(example_specs)
    print(f"\nInput check: {ok}\n  {msg}")

    if ok:
        matches = find_matches(example_specs, db, top_n=3)
        print_comparison(matches, example_specs)
